"""
AireLogix Underwriting Workbook Generator — v2.0
9-tab structure per v1.7 methodology
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import math
from datetime import datetime


# ── Brand colors (openpyxl uses ARGB hex) ────────────────────────────────────
NAVY      = "FF020C1A"
NAVY2     = "FF04101F"
NAVY3     = "FF0B1E3A"
GOLD      = "FFE8CC9A"
GOLD2     = "FFC8A96E"
GOLD3     = "FF9A7A14"
WHITE     = "FFFFFFFF"
STEEL     = "FF8BA4BE"
GREEN     = "FF2D9E5F"
RED       = "FFC0392B"
AMBER     = "FFE67E22"
LIGHT_BG  = "FFF5F0E8"
MID_BG    = "FFE8E0D0"

# Input (blue), formula (black), cross-sheet (green) per standard
BLUE_INPUT = "FF0000FF"
BLACK_CALC = "FF000000"
GREEN_LINK = "FF008000"


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color=BLACK_CALC, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _border(style="thin"):
    s = Side(style=style, color="FFD0C8B8")
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin", color="FFD0C8B8")
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _fmt_dollar(ws, cell_ref, decimals=0):
    ws[cell_ref].number_format = f'$#,##0{"." + "0"*decimals if decimals else ""};($#,##0{"." + "0"*decimals if decimals else ""});"-"'

def _fmt_pct(ws, cell_ref, decimals=1):
    ws[cell_ref].number_format = f'0.{"0"*decimals}%;-0.{"0"*decimals}%;"-"'

def _fmt_x(ws, cell_ref, decimals=2):
    ws[cell_ref].number_format = f'0.{"0"*decimals}x;-0.{"0"*decimals}x;"-"'

def _header_row(ws, row, title, col_span=8):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1, value=title)
    c.font = _font(bold=True, color=GOLD, size=10)
    c.fill = _fill(NAVY3)
    c.alignment = _align("left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)

def _label(ws, row, col, text, bold=False, indent=0):
    prefix = "  " * indent
    c = ws.cell(row=row, column=col, value=prefix + text)
    c.font = _font(bold=bold, color="FF2C3E50", size=9)
    c.alignment = _align("left")
    c.fill = _fill(LIGHT_BG)
    return c

def _value(ws, row, col, val, fmt="dollar", bold=False, color=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold=bold, color=color or BLACK_CALC, size=9)
    c.alignment = _align("right")
    c.fill = _fill("FFFBF8F2")
    if fmt == "dollar":
        c.number_format = '$#,##0;($#,##0);"-"'
    elif fmt == "pct":
        c.number_format = '0.0%;-0.0%;"-"'
    elif fmt == "x":
        c.number_format = '0.00x;-0.00x;"-"'
    elif fmt == "bps":
        c.number_format = '0" bps"'
    elif fmt == "text":
        c.alignment = _align("left")
        c.fill = _fill("FFFBF8F2")
    elif fmt == "int":
        c.number_format = '#,##0;-#,##0;"-"'
    return c

def _row(ws, row, label, val, fmt="dollar", indent=0, bold=False, val_color=None):
    _label(ws, row, 1, label, bold=bold, indent=indent)
    _value(ws, row, 2, val, fmt=fmt, bold=bold, color=val_color)
    ws.row_dimensions[row].height = 16

def _spacer(ws, row, h=6):
    ws.row_dimensions[row].height = h

def _setup_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── TAB 1: Dashboard ──────────────────────────────────────────────────────────
def _tab_dashboard(wb, a):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "9A7A14"

    tx  = a.get("transaction", {})
    rr  = a.get("riskRating", {})
    gd  = a.get("gdscr", {})
    bs  = a.get("balanceSheet", {})
    ac  = a.get("aircraft", {})
    col = a.get("collateral", {})

    _setup_cols(ws, [32, 20, 4, 32, 20])

    # Title block
    ws.row_dimensions[1].height = 30
    c = ws.cell(row=1, column=1, value="AIRELOGIX UNDERWRITING WORKBOOK")
    c.font = _font(bold=True, color=GOLD, size=14)
    c.fill = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.merge_cells("A1:E1")

    ws.row_dimensions[2].height = 18
    c = ws.cell(row=2, column=1, value=f"Confidential — {a.get('borrowerName','')} — {a.get('analysisDate','')}")
    c.font = _font(italic=True, color=STEEL, size=9)
    c.fill = _fill(NAVY2)
    c.alignment = _align("left", "center")
    ws.merge_cells("A2:E2")

    _spacer(ws, 3, 10)

    # Left column — Credit Summary
    r = 4
    _header_row(ws, r, "CREDIT SUMMARY", 2); r+=1
    _row(ws, r, "Borrower Name", a.get("borrowerName",""), "text", bold=True); r+=1
    _row(ws, r, "Application ID", a.get("applicationId",""), "text"); r+=1
    _row(ws, r, "Analysis Date", a.get("analysisDate",""), "text"); r+=1
    _spacer(ws, r); r+=1
    _row(ws, r, "AireLogix Risk Rating", str(rr.get("rating","")) + " — " + rr.get("band",""), "text", bold=True,
         val_color=GREEN if str(rr.get("rating","")) in ["1","2","3"] else AMBER if str(rr.get("rating","")) in ["4","5","6"] else RED); r+=1
    _row(ws, r, "Disposition", rr.get("disposition",""), "text"); r+=1
    _spacer(ws, r); r+=1
    _row(ws, r, "Pro Forma GDSCR", gd.get("gdscr",0), "x", bold=True,
         val_color=GREEN if gd.get("gdscr",0)>=1.75 else AMBER if gd.get("gdscr",0)>=1.25 else RED); r+=1
    _row(ws, r, "Net Worth Coverage", bs.get("netWorthCoverage",0), "x"); r+=1
    _row(ws, r, "Liquidity Ratio", bs.get("liquidityRatio",0), "x"); r+=1
    _spacer(ws, r); r+=1

    # Right column — Transaction
    r2 = 4
    # Right column header (cols D-E)
    c_rh = ws.cell(row=r2, column=4, value="TRANSACTION")
    c_rh.font = _font(bold=True, color=GOLD, size=10)
    c_rh.fill = _fill(NAVY3)
    c_rh.alignment = _align("left")
    ws.merge_cells(start_row=r2, start_column=4, end_row=r2, end_column=5)
    ws.row_dimensions[r2].height = 22
    # We'll put right column in columns D-E
    def _row_r(row, label, val, fmt="dollar", bold=False):
        c1 = ws.cell(row=row, column=4, value=label)
        c1.font = _font(bold=bold, color="FF2C3E50", size=9)
        c1.fill = _fill(LIGHT_BG)
        c1.alignment = _align("left")
        c2 = ws.cell(row=row, column=5, value=val)
        c2.font = _font(bold=bold, size=9)
        c2.fill = _fill("FFFBF8F2")
        c2.alignment = _align("right")
        ws.row_dimensions[row].height = 16
        if fmt == "dollar": c2.number_format = '$#,##0;($#,##0);"-"'
        elif fmt == "pct":  c2.number_format = '0.0%;-0.0%;"-"'
        elif fmt == "x":    c2.number_format = '0.00x;-0.00x;"-"'
        elif fmt == "text": c2.alignment = _align("left")

    r2 += 1
    _row_r(r2, "Aircraft", ac.get("description",""), "text", True); r2+=1
    _row_r(r2, "Serial Number", ac.get("serialNumber",""), "text"); r2+=1
    _row_r(r2, "Registration", ac.get("registration","N/A"), "text"); r2+=1
    _row_r(r2, "Engine Program", ac.get("engineProgram",""), "text"); r2+=1
    _row_r(r2, "Airframe Total Time", ac.get("aftt",0), "int"); r2+=1
    _spacer(ws, r2); r2+=1
    _row_r(r2, "Purchase Price", tx.get("purchasePrice",0), "dollar", True); r2+=1
    _row_r(r2, "Loan Amount", tx.get("loanAmount",0), "dollar"); r2+=1
    _row_r(r2, "LTV (vs Purchase Price)", tx.get("ltv",0), "pct"); r2+=1
    _row_r(r2, "LTV (vs AireLogix FMV)", tx.get("ltvVsFMV",0), "pct"); r2+=1
    _row_r(r2, "AireLogix Est. FMV", col.get("fmv_curve",0), "dollar"); r2+=1
    _row_r(r2, "Illustrative Rate", tx.get("illustrativeRate",0), "pct"); r2+=1
    _row_r(r2, "Term (Months)", tx.get("termMonths",0), "int"); r2+=1
    _row_r(r2, "Est. Monthly Payment", tx.get("monthlyPayment",0), "dollar", True); r2+=1

    # Flags section
    flags = a.get("flags", [])
    if flags:
        r = max(r, r2) + 2
        _header_row(ws, r, f"FLAGS & EXCEPTIONS ({len(flags)} flagged)", 5); r+=1
        for f in flags:
            sev = f.get("severity","")
            sev_color = RED if sev=="CRITICAL" else AMBER if sev=="MATERIAL" else STEEL
            c = ws.cell(row=r, column=1, value=f"  [{sev}] {f.get('title','')}")
            c.font = _font(color=sev_color, size=9)
            c.fill = _fill(LIGHT_BG)
            c.alignment = _align("left")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 16
            r+=1


# ── TAB 2: Income ─────────────────────────────────────────────────────────────
def _tab_income(wb, a):
    ws = wb.create_sheet("Income")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2D9E5F"
    _setup_cols(ws, [36, 18, 18, 18, 18])

    inc = a.get("incomeNormalization", {})
    years = inc.get("taxYears", [2023, 2024])
    y1 = years[0] if len(years) > 0 else "Year 1"
    y2 = years[1] if len(years) > 1 else "Year 2"
    gov_yr = inc.get("qualifyingYear") or inc.get("governingYear") or y1

    r = 1
    c = ws.cell(row=r, column=1, value="INCOME NORMALIZATION")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 26; r+=1

    c = ws.cell(row=r, column=1, value=inc.get("note",""))
    c.font = _font(italic=True, color=STEEL, size=8)
    c.fill = _fill(NAVY2)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 16; r+=1
    _spacer(ws, r); r+=1

    # Column headers
    _header_row(ws, r, "INCOME DETAIL", 5); r+=1
    for col_i, label in enumerate(["Source", str(y1), str(y2), "Qualifying", "Notes"], 1):
        c = ws.cell(row=r, column=col_i, value=label)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY3)
        c.alignment = _align("center" if col_i > 1 else "left")
    ws.row_dimensions[r].height = 18; r+=1

    # W2
    for w2 in inc.get("w2Detail", []):
        ws.cell(row=r, column=1, value="  W-2 Wages — " + w2.get("employer","")).font = _font(size=9)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        _value(ws, r, 2, w2.get("year1",0))
        _value(ws, r, 3, w2.get("year2",0))
        _value(ws, r, 4, min(w2.get("year1",0), w2.get("year2",0)))
        ws.row_dimensions[r].height = 16; r+=1

    # K-1s
    for k1 in inc.get("k1Detail", []):
        excluded = k1.get("excluded", False)
        color = STEEL if excluded else BLACK_CALC
        ws.cell(row=r, column=1, value="  K-1 — " + k1.get("entityName","")).font = _font(size=9, color=color)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        _value(ws, r, 2, k1.get("year1Box1",0) or k1.get("qualifyingIncome",0))
        _value(ws, r, 3, k1.get("year2Box1",0) or k1.get("qualifyingIncome",0))
        qi = 0 if excluded else k1.get("qualifyingIncome",0)
        _value(ws, r, 4, qi, color=STEEL if excluded else BLACK_CALC)
        note = k1.get("exclusionReason") or k1.get("note") or ("Excluded — LP passive" if excluded else k1.get("participationType",""))
        ws.cell(row=r, column=5, value=note).font = _font(size=8, italic=True, color=STEEL)
        ws.cell(row=r, column=5).fill = _fill("FFFBF8F2")
        ws.row_dimensions[r].height = 16; r+=1

    # Portfolio income
    pi = inc.get("portfolioIncome", {})
    if pi:
        for k, v in pi.items():
            if v:
                ws.cell(row=r, column=1, value=f"  {k.title()}").font = _font(size=9)
                ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
                _value(ws, r, 4, v)
                ws.row_dimensions[r].height = 16; r+=1

    _spacer(ws, r); r+=1

    # Totals
    _header_row(ws, r, "QUALIFYING INCOME SUMMARY", 5); r+=1
    _row(ws, r, "Year 1 Gross Income", inc.get("year1Total",0), bold=False); r+=1
    _row(ws, r, "Year 2 Gross Income", inc.get("year2Total",0), bold=False); r+=1

    var = inc.get("variancePct",0)
    var_color = RED if abs(var) > 25 else AMBER if abs(var) > 15 else BLACK_CALC
    _row(ws, r, "Year-over-Year Variance", var/100 if var else 0, "pct", val_color=var_color); r+=1
    _row(ws, r, f"Governing Year ({gov_yr})", inc.get("qualifyingIncome",0), bold=True,
         val_color=GREEN); r+=1
    _row(ws, r, "Estimated Taxes (36% effective)", inc.get("taxesPaidLowerYear",0)); r+=1
    _row(ws, r, "After-Tax Qualifying Income", inc.get("afterTaxQualifyingIncome",0), bold=False); r+=1


# ── TAB 3: K-1 Detail ────────────────────────────────────────────────────────
def _tab_k1(wb, a):
    ws = wb.create_sheet("K-1 Detail")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8A96E"
    _setup_cols(ws, [30, 22, 14, 14, 16, 24])

    inc = a.get("incomeNormalization", {})
    k1s = inc.get("k1Detail", [])
    years = inc.get("taxYears", [2023, 2024])
    y1 = years[0] if len(years) > 0 else "Year 1"
    y2 = years[1] if len(years) > 1 else "Year 2"

    r = 1
    c = ws.cell(row=r, column=1, value="K-1 SCHEDULE DETAIL")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:F{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "PASS-THROUGH ENTITY ANALYSIS", 6); r+=1
    headers = ["Entity Name", "Participation Type", str(y1) + " Box 1", str(y2) + " Box 1", "Qualifying Income", "Notes"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col_i, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY3)
        c.alignment = _align("center" if col_i > 2 else "left")
    ws.row_dimensions[r].height = 18; r+=1

    for k1 in k1s:
        excluded = k1.get("excluded", False)
        bg = "FFF5F0E8" if not excluded else "FFF0F0F0"
        color = STEEL if excluded else BLACK_CALC

        cells = [
            k1.get("entityName",""),
            k1.get("participationType",""),
            k1.get("year1Box1",0) or 0,
            k1.get("year2Box1",0) or 0,
            0 if excluded else (k1.get("qualifyingIncome",0) or 0),
            k1.get("exclusionReason") or k1.get("note","") or ("Excluded — LP passive" if excluded else "Active qualifying")
        ]
        fmts = ["text","text","dollar","dollar","dollar","text"]
        for col_i, (val, fmt) in enumerate(zip(cells, fmts), 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font = _font(size=9, color=color, italic=excluded)
            c.fill = _fill(bg)
            c.alignment = _align("right" if fmt == "dollar" else "left")
            if fmt == "dollar":
                c.number_format = '$#,##0;($#,##0);"-"'
        ws.row_dimensions[r].height = 16; r+=1

    if not k1s:
        c = ws.cell(row=r, column=1, value="No K-1 detail available — income sourced from W-2 or self-reported")
        c.font = _font(italic=True, color=STEEL, size=9)
        c.fill = _fill(LIGHT_BG)
        ws.merge_cells(f"A{r}:F{r}")
        ws.row_dimensions[r].height = 16; r+=1

    _spacer(ws, r); r+=1
    _header_row(ws, r, "METHODOLOGY NOTES", 6); r+=1
    notes = [
        "Lower of two tax years governs qualifying income (conservative principle)",
        "LP passive K-1 income excluded per v1.7 methodology",
        "Section 179 and depreciation add-backs not permitted",
        "Real estate K-1s: net income after depreciation only",
        "Capital gains excluded from qualifying income",
    ]
    for note in notes:
        c = ws.cell(row=r, column=1, value=f"  • {note}")
        c.font = _font(size=8, italic=True, color=STEEL)
        c.fill = _fill(LIGHT_BG)
        ws.merge_cells(f"A{r}:F{r}")
        ws.row_dimensions[r].height = 14; r+=1


# ── TAB 4: Debt Service ───────────────────────────────────────────────────────
def _tab_debt_service(wb, a):
    ws = wb.create_sheet("Debt Service")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0B1E3A"
    _setup_cols(ws, [36, 20, 20])

    tx  = a.get("transaction", {})
    gd  = a.get("gdscr", {})
    inc = a.get("incomeNormalization", {})

    r = 1
    c = ws.cell(row=r, column=1, value="PRO FORMA DEBT SERVICE ANALYSIS")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:C{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "PROPOSED AIRCRAFT FINANCING", 3); r+=1
    _row(ws, r, "Purchase Price", tx.get("purchasePrice",0)); r+=1
    _row(ws, r, "Loan Amount", tx.get("loanAmount",0), bold=True); r+=1
    _row(ws, r, "Down Payment", tx.get("purchasePrice",0) - tx.get("loanAmount",0)); r+=1
    _row(ws, r, "LTV (vs Purchase Price)", tx.get("ltv",0), "pct"); r+=1
    _row(ws, r, "LTV (vs AireLogix FMV)", tx.get("ltvVsFMV",0), "pct"); r+=1
    _spacer(ws, r); r+=1
    _row(ws, r, "Illustrative Rate", tx.get("illustrativeRate",0), "pct"); r+=1
    _row(ws, r, "Rate Basis", tx.get("rateNote",""), "text"); r+=1
    _row(ws, r, "Term (Months)", tx.get("termMonths",0), "int"); r+=1
    _row(ws, r, "Est. Monthly Payment (P&I)", tx.get("monthlyPayment",0), bold=True); r+=1
    _row(ws, r, "Annual Aircraft Debt Service", tx.get("annualAircraftDS",0), bold=True); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "EXISTING OBLIGATIONS", 3); r+=1
    _row(ws, r, "Existing Annual Debt Service", tx.get("existingAnnualDS",0), bold=True); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "PRO FORMA GLOBAL DEBT SERVICE", 3); r+=1
    _row(ws, r, "Aircraft DS (New)", tx.get("annualAircraftDS",0)); r+=1
    _row(ws, r, "Existing DS", tx.get("existingAnnualDS",0)); r+=1
    _row(ws, r, "Total Pro Forma DS", tx.get("totalProFormaDS",0), bold=True,
         val_color=BLACK_CALC); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "GDSCR CALCULATION", 3); r+=1
    _row(ws, r, "Gross Qualifying Income", inc.get("qualifyingIncome",0), bold=True); r+=1
    _row(ws, r, "Total Pro Forma DS", tx.get("totalProFormaDS",0)); r+=1
    gdscr = gd.get("gdscr",0)
    gdscr_color = GREEN if gdscr >= 1.75 else AMBER if gdscr >= 1.25 else RED
    _row(ws, r, "Pro Forma GDSCR", gdscr, "x", bold=True, val_color=gdscr_color); r+=1
    _row(ws, r, "Assessment", gd.get("assessment",""), "text"); r+=1
    _spacer(ws, r); r+=1

    # Balloon projection
    _header_row(ws, r, "BALLOON PROJECTION (10-YEAR)", 3); r+=1
    fmv = a.get("collateral",{}).get("fmv_curve",0) or a.get("transaction",{}).get("fmv",0)
    depr = 0.030
    for yr in [5, 7, 10]:
        fmv_at = fmv * ((1-depr)**yr)
        balloon = fmv_at * 0.70
        _row(ws, r, f"Projected FMV at Year {yr}", fmv_at); r+=1
        _row(ws, r, f"  Balloon (70% of FMV)", balloon, indent=1); r+=1


# ── TAB 5: Balance Sheet ──────────────────────────────────────────────────────
def _tab_balance_sheet(wb, a):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2D9E5F"
    _setup_cols(ws, [36, 20, 20])

    bs = a.get("balanceSheet", {})

    r = 1
    c = ws.cell(row=r, column=1, value="PERSONAL BALANCE SHEET")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:C{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "ASSETS", 3); r+=1
    _row(ws, r, "Gross Total Assets", bs.get("grossTotalAssets",0), bold=True); r+=1
    _row(ws, r, "  Gross Liquid Assets (Tier 1)", bs.get("tier1GrossLiquid",0), indent=1); r+=1
    _row(ws, r, "  Margin Loan Adjustment", -(bs.get("marginLoanAdjustment",0) or 0), indent=1); r+=1
    _row(ws, r, "  Net Liquid Assets", bs.get("tier1NetLiquid",0), indent=1, bold=True); r+=1

    # Tier 1 detail
    for asset in bs.get("tier1Assets", []):
        if isinstance(asset, dict):
            _row(ws, r, f"    {asset.get('description','Asset')}", asset.get("net",0), indent=2); r+=1

    _spacer(ws, r); r+=1
    _header_row(ws, r, "LIABILITIES", 3); r+=1
    _row(ws, r, "Total Liabilities", bs.get("totalLiabilities",0), bold=True); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "NET WORTH", 3); r+=1
    nw = bs.get("statedNetWorth",0)
    _row(ws, r, "Stated Net Worth", nw, bold=True,
         val_color=GREEN if nw > 0 else RED); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "LIQUIDITY RATIOS", 3); r+=1
    liq = bs.get("liquidityRatio",0)
    liq_color = GREEN if liq >= 2.0 else AMBER if liq >= 1.0 else RED
    _row(ws, r, "Liquidity Ratio (Liquid / Loan)", liq, "x", val_color=liq_color); r+=1
    _row(ws, r, "Liquidity Assessment", bs.get("liquidityAssessment",""), "text"); r+=1
    _row(ws, r, "Leverage Ratio (Liab / Assets)", bs.get("leverageRatio",0), "x"); r+=1

    # Entity guarantee
    egl = bs.get("entityGuaranteeLiquidity",{})
    if egl.get("adjustmentApplied"):
        _spacer(ws, r); r+=1
        _header_row(ws, r, "ENTITY GUARANTEE POOL", 3); r+=1
        _row(ws, r, "Base Personal Liquid", egl.get("baseLiquid",0)); r+=1
        _row(ws, r, "Entity Guarantee Adjustment", egl.get("guaranteeAdjustment",0)); r+=1
        _row(ws, r, "Adjusted Effective Liquid", egl.get("adjustedLiquid",0), bold=True); r+=1
        _row(ws, r, "Adjusted Liquidity Ratio", egl.get("adjustedLiquidityRatio",0), "x"); r+=1


# ── TAB 6: Net Worth & Leverage ───────────────────────────────────────────────
def _tab_nw_leverage(wb, a):
    ws = wb.create_sheet("Net Worth & Leverage")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "9A7A14"
    _setup_cols(ws, [36, 20, 20])

    bs  = a.get("balanceSheet", {})
    tx  = a.get("transaction", {})
    gd  = a.get("gdscr", {})
    loan = tx.get("loanAmount", 0)
    nw   = bs.get("statedNetWorth", 0)
    liq  = bs.get("tier1NetLiquid", 0)

    r = 1
    c = ws.cell(row=r, column=1, value="NET WORTH & LEVERAGE ANALYSIS")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:C{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "COVERAGE MULTIPLES", 3); r+=1
    nw_cov = nw/loan if loan else 0
    nw_color = GREEN if nw_cov >= 3.0 else AMBER if nw_cov >= 1.5 else RED
    _row(ws, r, "Net Worth Coverage (NW / Loan)", nw_cov, "x", bold=True, val_color=nw_color); r+=1
    liq_cov = liq/loan if loan else 0
    liq_color = GREEN if liq_cov >= 1.5 else AMBER if liq_cov >= 1.0 else RED
    _row(ws, r, "Liquidity Coverage (Liq / Loan)", liq_cov, "x", bold=True, val_color=liq_color); r+=1
    _row(ws, r, "GDSCR (Income / Total DS)", gd.get("gdscr",0), "x"); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "BALANCE SHEET COMPOSITION", 3); r+=1
    gross = bs.get("grossTotalAssets",0)
    liab  = bs.get("totalLiabilities",0)
    _row(ws, r, "Gross Total Assets", gross); r+=1
    _row(ws, r, "Total Liabilities", liab); r+=1
    _row(ws, r, "Stated Net Worth", nw, bold=True); r+=1
    _row(ws, r, "Leverage (Liab / Assets)", liab/gross if gross else 0, "x"); r+=1
    _row(ws, r, "Liquid as % of Total Assets", liq/gross if gross else 0, "pct"); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "GUARANTEE STRUCTURE", 3); r+=1
    guarantors = a.get("guarantors", [])
    for g in guarantors:
        _row(ws, r, g.get("name","") + " — " + g.get("type",""), "", "text", bold=True); r+=1
        _row(ws, r, "  Net Worth", g.get("netWorth",0), indent=1); r+=1
        _row(ws, r, "  Liquid Assets", g.get("liquidAssets",0), indent=1); r+=1
        _row(ws, r, "  Guarantee Type", g.get("guaranteeType",""), "text", indent=1); r+=1
        _spacer(ws, r); r+=1

    _header_row(ws, r, "BENCHMARK THRESHOLDS", 3); r+=1
    benchmarks = [
        ("Net Worth Coverage", "≥ 3.0x", "Strong", "2.0-3.0x", "Adequate", "< 2.0x", "Weak"),
        ("Liquidity Coverage", "≥ 1.5x", "Strong", "1.0-1.5x", "Adequate", "< 1.0x", "Critical"),
        ("GDSCR",             "≥ 1.75x","Strong", "1.25-1.75x","Adequate","< 1.25x","Marginal"),
    ]
    for b in benchmarks:
        c = ws.cell(row=r, column=1, value=b[0])
        c.font = _font(bold=True, size=9)
        c.fill = _fill(LIGHT_BG)
        c2 = ws.cell(row=r, column=2, value=f"{b[1]} = {b[2]}  |  {b[3]} = {b[4]}  |  {b[5]} = {b[6]}")
        c2.font = _font(size=8, italic=True, color=STEEL)
        c2.fill = _fill("FFFBF8F2")
        ws.row_dimensions[r].height = 16; r+=1


# ── TAB 7: Collateral ─────────────────────────────────────────────────────────
def _tab_collateral(wb, a):
    ws = wb.create_sheet("Collateral")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8A96E"
    _setup_cols(ws, [36, 20, 20, 20])

    col = a.get("collateral", {})
    tx  = a.get("transaction", {})
    ac  = a.get("aircraft", {})
    fmv = col.get("fmv_curve", 0) or tx.get("fmv", 0)
    loan = tx.get("loanAmount", 0)

    r = 1
    c = ws.cell(row=r, column=1, value="COLLATERAL ANALYSIS")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:D{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "AIRCRAFT DESCRIPTION", 4); r+=1
    _row(ws, r, "Description", ac.get("description",""), "text", bold=True); r+=1
    _row(ws, r, "Serial Number", col.get("sn","") or ac.get("serialNumber",""), "text"); r+=1
    _row(ws, r, "Registration", col.get("registration","") or ac.get("registration","N/A"), "text"); r+=1
    _row(ws, r, "Model Year", col.get("year", ac.get("year",0)), "int"); r+=1
    _row(ws, r, "Airframe Total Time", col.get("aftt", ac.get("aftt",0)), "int"); r+=1
    _row(ws, r, "Engine Program", col.get("engine_program","") or ac.get("engineProgram",""), "text"); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "VALUATION", 4); r+=1
    _row(ws, r, "Curve Used", col.get("curveName","AireLogix Curve"), "text"); r+=1
    _row(ws, r, "AireLogix Est. FMV", fmv, bold=True,
         val_color=GREEN); r+=1
    _row(ws, r, "Purchase Price", col.get("fmv_purchase_price",0) or tx.get("purchasePrice",0)); r+=1
    pp_vs_fmv = (col.get("fmv_purchase_price",0) / fmv - 1) if fmv else 0
    pp_color = GREEN if abs(pp_vs_fmv) < 0.05 else AMBER if abs(pp_vs_fmv) < 0.15 else RED
    _row(ws, r, "Purchase Price vs FMV", pp_vs_fmv, "pct", val_color=pp_color); r+=1
    _row(ws, r, "Loan Amount", loan); r+=1
    _row(ws, r, "LTV vs AireLogix FMV", col.get("ltv_on_curve_fmv",0)/100, "pct", bold=True); r+=1
    _row(ws, r, "LTV vs Purchase Price", col.get("ltv_on_purchase_price",0)/100, "pct"); r+=1
    _spacer(ws, r); r+=1

    # Forward depreciation table
    _header_row(ws, r, "FORWARD DEPRECIATION SCHEDULE (Base 3.0% / Bear 5.0% annual)", 4); r+=1
    for col_i, h in enumerate(["Year", "Base FMV", "Bear FMV", "Loan Balance (Est.)"], 1):
        c = ws.cell(row=r, column=col_i, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY3)
        c.alignment = _align("center" if col_i > 1 else "left")
    ws.row_dimensions[r].height = 18; r+=1

    base_rate = 0.030
    bear_rate = 0.050
    mo_rate   = tx.get("illustrativeRate",0.063)/12
    mo_pmt    = tx.get("monthlyPayment",0)
    bal = loan

    for yr in range(0, 11):
        fmv_base = fmv * ((1-base_rate)**yr)
        fmv_bear = fmv * ((1-bear_rate)**yr)
        if yr > 0:
            for _ in range(12):
                interest = bal * mo_rate
                principal = max(mo_pmt - interest, 0)
                bal = max(bal - principal, 0)

        ws.cell(row=r, column=1, value=f"Year {yr}" if yr > 0 else "Now").font = _font(size=9)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        _value(ws, r, 2, round(fmv_base))
        _value(ws, r, 3, round(fmv_bear))
        _value(ws, r, 4, round(bal))
        ws.row_dimensions[r].height = 15; r+=1


# ── TAB 8: Risk Rating ────────────────────────────────────────────────────────
def _tab_risk_rating(wb, a):
    ws = wb.create_sheet("Risk Rating")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "E8CC9A"
    _setup_cols(ws, [36, 14, 14, 14, 22])

    rr = a.get("riskRating", {})
    fs = rr.get("factorScores", {})
    comp = rr.get("composite", {})

    r = 1
    c = ws.cell(row=r, column=1, value="AIRELOGIX RISK RATING (ORR)")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    rating = str(rr.get("rating",""))
    band   = rr.get("band","")
    rating_color = GREEN if rating in ["1","2","3"] else AMBER if rating in ["4","5","6"] else RED

    # Rating badge
    _header_row(ws, r, "FINAL RATING", 5); r+=1
    c = ws.cell(row=r, column=1, value=f"ORR {rating} — {band}")
    c.font = _font(bold=True, color=rating_color, size=14)
    c.fill = _fill(LIGHT_BG)
    ws.merge_cells(f"A{r}:B{r}")
    ws.row_dimensions[r].height = 24

    c2 = ws.cell(row=r, column=3, value=rr.get("disposition",""))
    c2.font = _font(size=9, italic=True, color=STEEL)
    c2.fill = _fill("FFFBF8F2")
    ws.merge_cells(f"C{r}:E{r}")
    r+=1
    _spacer(ws, r); r+=1

    # Factor scores
    _header_row(ws, r, "FACTOR SCORE DETAIL", 5); r+=1
    for col_i, h in enumerate(["Factor", "Score", "Weight", "Weighted", "Basis"], 1):
        c = ws.cell(row=r, column=col_i, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY3)
        c.alignment = _align("center" if col_i > 1 else "left")
    ws.row_dimensions[r].height = 18; r+=1

    factor_labels = {
        "F1_GDSCR": "F1 — Pro Forma GDSCR",
        "F2_Liquidity": "F2 — Liquidity",
        "F3_NetWorth": "F3 — Net Worth",
        "F4_IncomeQuality": "F4 — Income Quality",
        "F5_LTV": "F5 — LTV vs FMV",
        "F6_Collateral": "F6 — Collateral",
    }

    for key, label in factor_labels.items():
        f = fs.get(key, {})
        score = f.get("score", 0)
        score_color = GREEN if score <= 2 else AMBER if score <= 4 else RED if score <= 6 else "FFCC0000"
        bg = "FFFBF8F2"

        ws.cell(row=r, column=1, value=label).font = _font(size=9)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        c2 = ws.cell(row=r, column=2, value=score)
        c2.font = _font(bold=True, color=score_color, size=9)
        c2.fill = _fill(bg)
        c2.alignment = _align("center")
        c3 = ws.cell(row=r, column=3, value=f.get("weight",0))
        c3.number_format = '0%'
        c3.font = _font(size=9)
        c3.fill = _fill(bg)
        c3.alignment = _align("center")
        c4 = ws.cell(row=r, column=4, value=f.get("weighted",0))
        c4.number_format = '0.000'
        c4.font = _font(size=9)
        c4.fill = _fill(bg)
        c4.alignment = _align("center")
        ws.cell(row=r, column=5, value=f.get("basis","")).font = _font(size=8, italic=True, color=STEEL)
        ws.cell(row=r, column=5).fill = _fill(bg)
        ws.row_dimensions[r].height = 16; r+=1

    _spacer(ws, r); r+=1
    _header_row(ws, r, "COMPOSITE SCORE", 5); r+=1
    _row(ws, r, "Pre-Floor Composite", comp.get("preFloorComposite",0), "x"); r+=1
    if comp.get("floorTriggered"):
        _row(ws, r, "Floor Applied", comp.get("floorTriggered",""), "text", val_color=AMBER); r+=1
    _row(ws, r, "Final Composite Score", comp.get("finalComposite",0), "x", bold=True,
         val_color=rating_color); r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "ORR SCALE REFERENCE", 5); r+=1
    scale = [
        ("1", "Exceptional", "Approve — best execution"),
        ("2", "Very Strong", "Approve — full lender universe"),
        ("3", "Strong", "Approve — standard terms"),
        ("4", "Good", "Approve with conditions"),
        ("5", "Acceptable", "Approve with enhanced structure"),
        ("6- / 6+", "Marginal", "Present with full risk disclosure"),
        ("7", "Difficult", "Conditional placement — limited universe"),
        ("8", "Hard Decline", "Do not proceed"),
    ]
    for orr, band_s, action in scale:
        is_current = str(rr.get("rating","")) in orr
        bg = NAVY3 if is_current else LIGHT_BG
        fg = GOLD if is_current else STEEL
        for col_i, val in enumerate([orr, band_s, action], 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font = _font(bold=is_current, color=fg, size=9)
            c.fill = _fill(bg)
        ws.row_dimensions[r].height = 15; r+=1


# ── TAB 9: Trend Analysis ─────────────────────────────────────────────────────
def _tab_trend(wb, a):
    ws = wb.create_sheet("Trend Analysis")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4A6480"
    _setup_cols(ws, [36, 18, 18, 18, 18])

    inc = a.get("incomeNormalization", {})
    bs  = a.get("balanceSheet", {})
    tx  = a.get("transaction", {})
    years = inc.get("taxYears", [2023, 2024])
    y1 = years[0] if len(years) > 0 else "Year 1"
    y2 = years[1] if len(years) > 1 else "Year 2"
    gov_yr = inc.get("qualifyingYear") or inc.get("governingYear") or y1

    r = 1
    c = ws.cell(row=r, column=1, value="TREND ANALYSIS")
    c.font = _font(bold=True, color=GOLD, size=12)
    c.fill = _fill(NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 26; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "INCOME TREND", 5); r+=1
    for col_i, h in enumerate(["Metric", str(y1), str(y2), "Change", "Direction"], 1):
        c = ws.cell(row=r, column=col_i, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY3)
        c.alignment = _align("center" if col_i > 1 else "left")
    ws.row_dimensions[r].height = 18; r+=1

    y1_inc = inc.get("year1Total", 0)
    y2_inc = inc.get("year2Total", 0)
    change = y2_inc - y1_inc
    pct    = change/y1_inc if y1_inc else 0
    direction = "▲ Improving" if change > 0 else "▼ Declining" if change < 0 else "— Flat"
    dir_color = GREEN if change > 0 else RED if change < 0 else STEEL

    def _trend_row(label, v1, v2, fmt="dollar"):
        nonlocal r
        chg = v2 - v1
        pct_chg = chg/v1 if v1 else 0
        dir_s = "▲" if chg > 0 else "▼" if chg < 0 else "—"
        d_color = GREEN if chg > 0 else RED if chg < 0 else STEEL

        ws.cell(row=r, column=1, value=label).font = _font(size=9)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        _value(ws, r, 2, v1, fmt)
        _value(ws, r, 3, v2, fmt)
        _value(ws, r, 4, chg, fmt)
        c5 = ws.cell(row=r, column=5, value=f"{dir_s} {abs(pct_chg)*100:.1f}%")
        c5.font = _font(size=9, color=d_color, bold=True)
        c5.fill = _fill("FFFBF8F2")
        c5.alignment = _align("center")
        ws.row_dimensions[r].height = 16; r+=1

    _trend_row("Total Qualifying Income", y1_inc, y2_inc)

    for k1 in inc.get("k1Detail", []):
        if not k1.get("excluded"):
            _trend_row(f"  K-1: {k1.get('entityName','')[:25]}",
                       k1.get("year1Box1",0) or 0,
                       k1.get("year2Box1",0) or 0)

    _spacer(ws, r); r+=1
    _header_row(ws, r, "INCOME VARIANCE ASSESSMENT", 5); r+=1
    var = inc.get("variancePct", 0)
    var_label = "Within tolerance (≤15%)" if abs(var) <= 15 else "Elevated variance (15–25%)" if abs(var) <= 25 else "Material variance (>25%) — flag"
    var_color = GREEN if abs(var) <= 15 else AMBER if abs(var) <= 25 else RED
    _row(ws, r, "Year-over-Year Variance", var/100 if var else 0, "pct", val_color=var_color); r+=1

    c = ws.cell(row=r, column=1, value=f"  Assessment: {var_label}")
    c.font = _font(italic=True, color=var_color, size=9)
    c.fill = _fill(LIGHT_BG)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 16; r+=1
    _spacer(ws, r); r+=1

    _header_row(ws, r, "KEY RATIO BENCHMARKS", 5); r+=1
    gdscr = a.get("gdscr",{}).get("gdscr",0)
    nw_cov = bs.get("netWorthCoverage",0)
    liq_rat = bs.get("liquidityRatio",0)

    ratios = [
        ("Pro Forma GDSCR", gdscr, 1.75, 1.25, "x"),
        ("Net Worth Coverage", nw_cov, 3.0, 1.5, "x"),
        ("Liquidity Ratio", liq_rat, 2.0, 1.0, "x"),
    ]
    for label, val, strong_thresh, ok_thresh, fmt in ratios:
        color = GREEN if val >= strong_thresh else AMBER if val >= ok_thresh else RED
        benchmark = f"Strong ≥{strong_thresh}{fmt}  |  Adequate ≥{ok_thresh}{fmt}  |  Weak below"
        ws.cell(row=r, column=1, value=label).font = _font(bold=True, size=9)
        ws.cell(row=r, column=1).fill = _fill(LIGHT_BG)
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = _font(bold=True, color=color, size=11)
        c2.fill = _fill("FFFBF8F2")
        c2.alignment = _align("center")
        c2.number_format = f'0.00"{fmt}"'
        c3 = ws.cell(row=r, column=3, value=benchmark)
        c3.font = _font(size=8, italic=True, color=STEEL)
        c3.fill = _fill("FFFBF8F2")
        ws.merge_cells(f"C{r}:E{r}")
        ws.row_dimensions[r].height = 18; r+=1


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────
def generate_workbook(analysis: dict, output_path: str):
    wb = Workbook()

    _tab_dashboard(wb, analysis)
    _tab_income(wb, analysis)
    _tab_k1(wb, analysis)
    _tab_debt_service(wb, analysis)
    _tab_balance_sheet(wb, analysis)
    _tab_nw_leverage(wb, analysis)
    _tab_collateral(wb, analysis)
    _tab_risk_rating(wb, analysis)
    _tab_trend(wb, analysis)

    wb.save(output_path)
    return output_path
